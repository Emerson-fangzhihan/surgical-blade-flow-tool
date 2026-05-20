"""
流向数据分析工具 — 珠海海纳融成
支持上传多个月份流向 Excel，自动识别手术刀数据，生成可视化报表
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import io, re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# 页面配置
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="流向分析工具 | 海纳融成",
    page_icon="🔪",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─────────────────────────────────────────────
# 自定义 CSS
# ─────────────────────────────────────────────
st.markdown("""
<style>
  .block-container { padding-top: 1.5rem; }
  .metric-card {
    background: linear-gradient(135deg, #1F4E79 0%, #2E6DA4 100%);
    color: white;
    border-radius: 12px;
    padding: 1rem 1.2rem;
    text-align: center;
  }
  .metric-card .num { font-size: 2rem; font-weight: 700; }
  .metric-card .lbl { font-size: 0.8rem; opacity: 0.85; margin-top: 4px; }
  .stDataFrame { border-radius: 8px; }
  div[data-testid="stMetric"] {
    background: #f5f9ff;
    border: 1px solid #cde;
    border-radius: 10px;
    padding: 12px 16px;
  }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# 月份排序辅助函数
# ─────────────────────────────────────────────
MONTH_ORDER = {
    "1月": "2025-01", "2月": "2025-02", "3月": "2025-03",
    "4月": "2025-04", "5月": "2025-05", "6月": "2025-06",
    "7月": "2025-07", "8月": "2025-08", "9月": "2025-09",
    "10月": "2025-10", "11月": "2024-11", "12月": "2024-12",
}

def parse_month_label(fname: str) -> str:
    """从文件名推断月份标签，例如 '11月' / '1月'"""
    m = re.search(r'(\d{1,2})月', fname)
    if m:
        return f"{m.group(1)}月"
    return fname

# ─────────────────────────────────────────────
# 智能列检测：从 DataFrame 自动识别字段
# ─────────────────────────────────────────────
KEYWORD_MAPS = {
    "customer": ["购货单位", "客户", "医院", "机构", "单位名称"],
    "product":  ["货品名称", "品名", "产品名称", "商品名称"],
    "model":    ["规格型号", "型号", "规格", "款式"],
    "qty":      ["数量", "销售数量", "出库数量", "实发数量"],
}

def detect_col(df: pd.DataFrame, role: str) -> int | None:
    """返回列索引，找不到返回 None"""
    keywords = KEYWORD_MAPS[role]
    for i, col in enumerate(df.columns):
        col_str = str(col).strip()
        for kw in keywords:
            if kw in col_str:
                return i
    # 用列0~10内容关键词扫描
    sample_rows = min(5, len(df))
    for i, col in enumerate(df.columns):
        cell_vals = df.iloc[:sample_rows, i].astype(str).str.cat(sep=" ")
        for kw in keywords:
            if kw in cell_vals:
                return i
    return None

def auto_detect_header(xl_bytes: bytes, sheet_name: str) -> int:
    """自动检测表头所在行（0-based）"""
    try:
        df_raw = pd.read_excel(io.BytesIO(xl_bytes), sheet_name=sheet_name, header=None, nrows=15)
        for i, row in df_raw.iterrows():
            try:
                row_str = " ".join(str(x) if pd.notna(x) else "" for x in row)
                if any(kw in row_str for kw in ["货品名称", "品名", "购货单位", "客户", "单位名称"]):
                    return i
            except Exception:
                continue
    except Exception:
        pass
    return 0

# ─────────────────────────────────────────────
# 读取单个 Excel 文件 → 返回手术刀 DataFrame
# ─────────────────────────────────────────────
def load_one_file(uploaded_file, month_label: str) -> pd.DataFrame | None:
    xl_bytes = uploaded_file.read()
    try:
        xl = pd.ExcelFile(io.BytesIO(xl_bytes))
    except Exception as e:
        st.warning(f"⚠️ 无法读取 {uploaded_file.name}：{e}")
        return None

    # 选 Sheet：优先包含"明细"或"流向"的 Sheet
    sheet_name = xl.sheet_names[0]
    for s in xl.sheet_names:
        if any(kw in s for kw in ["明细", "流向", "数据"]):
            sheet_name = s
            break

    header_row = auto_detect_header(xl_bytes, sheet_name)
    df = pd.read_excel(io.BytesIO(xl_bytes), sheet_name=sheet_name, header=header_row)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(how="all")

    # 检测列
    ci = {role: detect_col(df, role) for role in ["customer", "product", "model", "qty"]}
    missing = [r for r, idx in ci.items() if idx is None]
    if missing:
        # 用列索引兜底（按过往经验）
        ci_fallback = {"customer": 1, "product": 2, "model": 3, "qty": 8}
        for r in missing:
            ci[r] = ci_fallback.get(r)
        st.info(f"ℹ️ {uploaded_file.name}：部分列自动识别失败({missing})，已按默认索引填充，请在下方确认数据正确。")

    # 取列
    try:
        cols = list(df.columns)
        result = pd.DataFrame({
            "客户": df.iloc[:, ci["customer"]],
            "货品名称": df.iloc[:, ci["product"]],
            "型号": df.iloc[:, ci["model"]],
            "数量": pd.to_numeric(df.iloc[:, ci["qty"]], errors="coerce"),
        })
    except Exception as e:
        st.warning(f"⚠️ {uploaded_file.name} 列提取失败：{e}")
        return None

    result = result.dropna(subset=["货品名称"])
    # 过滤手术刀
    mask = result["货品名称"].astype(str).str.contains("手术刀", na=False)
    result = result[mask].copy()
    result["月份"] = month_label
    result["数量"] = result["数量"].fillna(0).astype(int)
    return result if len(result) > 0 else None

# ─────────────────────────────────────────────
# 生成 Excel 透视表 (带热力色彩)
# ─────────────────────────────────────────────
def build_excel(pivot: pd.DataFrame, detail: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pivot.to_excel(writer, sheet_name="客户月度透视", index=True)
        detail.to_excel(writer, sheet_name="明细数据", index=False)

        wb = writer.book
        ws = writer.sheets["客户月度透视"]

        # 热力色：按列最大值着色
        from openpyxl.styles.fills import GradientFill
        fills_low  = PatternFill("solid", fgColor="FFFFFF")
        fills_high = PatternFill("solid", fgColor="BDD7EE")

        max_row = ws.max_row
        max_col = ws.max_column
        for col_idx in range(2, max_col + 1):
            vals = []
            for row_idx in range(2, max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    vals.append(float(cell.value) if cell.value else 0)
                except:
                    vals.append(0)
            col_max = max(vals) or 1
            for row_idx in range(2, max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    ratio = float(cell.value) / col_max if cell.value else 0
                except:
                    ratio = 0
                r = int(189 + (255 - 189) * (1 - ratio))
                g = int(215 + (255 - 215) * (1 - ratio))
                b = int(238 + (255 - 238) * (1 - ratio))
                hex_color = f"{r:02X}{g:02X}{b:02X}"
                cell.fill = PatternFill("solid", fgColor=hex_color)

        # 表头样式
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # 自适应列宽
        for col in ws.columns:
            max_w = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_w + 4, 30)

    return output.getvalue()

# ─────────────────────────────────────────────
# ══════════════ 主界面 ══════════════
# ─────────────────────────────────────────────
st.title("🔪 流向数据分析工具")
st.caption("上传代理商每月流向 Excel → 自动生成手术刀销售报表")

# ─── 侧边栏 ───
with st.sidebar:
    st.header("📂 上传流向文件")
    st.markdown("""
**使用说明：**
1. 每次可上传 1~12 个月的流向 Excel
2. 文件名需包含月份，例如 `11月流向.xlsx`
3. 支持多 Sheet，自动识别含"明细"/"流向"的表
4. 系统自动筛选**手术刀**产品

**列自动识别：** 购货单位、货品名称、规格型号、数量
    """)
    st.divider()
    uploaded_files = st.file_uploader(
        "选择 Excel 文件（可多选）",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
        help="同时选中多个文件上传"
    )
    if uploaded_files:
        st.success(f"已选择 {len(uploaded_files)} 个文件")
    
    st.divider()
    st.markdown("**筛选选项**")
    filter_zero = st.checkbox("隐藏数量为0的记录", value=True)
    top_n = st.slider("只显示前N名客户", 5, 20, 10)

# ─── 无文件时显示引导 ───
if not uploaded_files:
    col1, col2, col3 = st.columns(3)
    with col1:
        st.info("**第一步**\n\n点击左侧 「选择Excel文件」，上传1到多个月的流向Excel")
    with col2:
        st.info("**第二步**\n\n系统自动识别手术刀数据，无需手动配置列")
    with col3:
        st.info("**第三步**\n\n查看KPI、折线图、饼图、透视表，下载Excel报表")
    
    st.markdown("---")
    st.markdown("### 📋 支持的文件格式示例")
    st.markdown("""
| 文件名示例 | Sheet | 自动识别列 |
|-----------|-------|----------|
| `11月流向.xlsx` | 11月明细 | 购货单位、货品名称、规格型号、数量 |
| `2月珠海海纳融成.xlsx` | 2月明细 | 自动匹配 |
| `3月流向1.xlsx` | 3月流向 | 自动匹配 |
    """)
    st.stop()

# ─── 读取并合并所有文件 ───
all_dfs = []
progress = st.progress(0, text="正在读取文件…")
for i, f in enumerate(uploaded_files):
    month_label = parse_month_label(f.name)
    df_one = load_one_file(f, month_label)
    if df_one is not None and len(df_one) > 0:
        all_dfs.append(df_one)
    progress.progress((i + 1) / len(uploaded_files), text=f"读取中: {f.name}")
progress.empty()

if not all_dfs:
    st.error("⚠️ 所有文件均未找到手术刀数据，请检查文件格式或联系技术支持。")
    st.stop()

df = pd.concat(all_dfs, ignore_index=True)
if filter_zero:
    df = df[df["数量"] > 0]

# 月份排序
months_in_data = sorted(df["月份"].unique(),
                        key=lambda m: MONTH_ORDER.get(m, m))

# ─────────────────────────────────────────────
# KPI 看板
# ─────────────────────────────────────────────
st.markdown("## 📊 总览指标")
total_qty  = int(df["数量"].sum())
total_cust = df["客户"].nunique()
total_model = df["型号"].nunique()
top_cust   = df.groupby("客户")["数量"].sum().idxmax()
top_cust_qty = int(df.groupby("客户")["数量"].sum().max())
best_month = df.groupby("月份")["数量"].sum()
best_month_label = best_month.idxmax() if len(best_month) > 0 else "-"
best_month_qty   = int(best_month.max()) if len(best_month) > 0 else 0

c1, c2, c3, c4, c5 = st.columns(5)
c1.metric("总销量（把）", f"{total_qty:,}")
c2.metric("客户数", total_cust)
c3.metric("型号数", total_model)
c4.metric("销量最高月", best_month_label, f"{best_month_qty:,} 把")
c5.metric("头部客户", top_cust[:8] + "…" if len(top_cust) > 8 else top_cust, f"{top_cust_qty:,} 把")

st.markdown("---")

# ─────────────────────────────────────────────
# Tab 布局
# ─────────────────────────────────────────────
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📈 月度趋势", "🥧 客户占比", "📦 月度结构", "📋 客户透视表", "🔍 明细数据"
])

# ─── Tab1：月度趋势折线图 ───
with tab1:
    st.markdown("### 客户月度销量趋势（TOP N）")
    cust_total = df.groupby("客户")["数量"].sum().sort_values(ascending=False)
    top_custs = cust_total.head(top_n).index.tolist()
    df_top = df[df["客户"].isin(top_custs)]
    
    trend = df_top.groupby(["月份", "客户"])["数量"].sum().reset_index()
    trend["月份排序"] = trend["月份"].map(lambda m: MONTH_ORDER.get(m, m))
    trend = trend.sort_values("月份排序")

    fig_line = px.line(
        trend, x="月份", y="数量", color="客户",
        markers=True,
        color_discrete_sequence=px.colors.qualitative.Set2,
        title=f"TOP{top_n} 客户月度趋势"
    )
    fig_line.update_layout(
        height=420,
        plot_bgcolor="white",
        paper_bgcolor="white",
        legend=dict(orientation="h", y=-0.2),
        xaxis=dict(categoryorder="array", categoryarray=months_in_data)
    )
    fig_line.update_traces(line_width=2, marker_size=7)
    st.plotly_chart(fig_line, use_container_width=True)

    # 月度合计折线
    monthly_total = df.groupby("月份")["数量"].sum().reset_index()
    monthly_total["月份排序"] = monthly_total["月份"].map(lambda m: MONTH_ORDER.get(m, m))
    monthly_total = monthly_total.sort_values("月份排序")
    fig_total = px.bar(
        monthly_total, x="月份", y="数量",
        title="各月总销量",
        color_discrete_sequence=["#2E6DA4"],
        text="数量"
    )
    fig_total.update_traces(texttemplate="%{text:,}", textposition="outside")
    fig_total.update_layout(
        height=300, plot_bgcolor="white", paper_bgcolor="white",
        xaxis=dict(categoryorder="array", categoryarray=months_in_data)
    )
    st.plotly_chart(fig_total, use_container_width=True)

# ─── Tab2：客户占比饼图 ───
with tab2:
    st.markdown("### 客户销量占比")
    cust_sum = df.groupby("客户")["数量"].sum().sort_values(ascending=False).reset_index()
    cust_sum.columns = ["客户", "总数量"]
    
    # 超出 top_n 合并为"其他"
    if len(cust_sum) > top_n:
        top_df = cust_sum.head(top_n)
        other_qty = cust_sum.tail(len(cust_sum) - top_n)["总数量"].sum()
        other_row = pd.DataFrame([{"客户": f"其他({len(cust_sum)-top_n}家)", "总数量": other_qty}])
        cust_sum_plot = pd.concat([top_df, other_row], ignore_index=True)
    else:
        cust_sum_plot = cust_sum.copy()

    fig_pie = px.pie(
        cust_sum_plot, names="客户", values="总数量",
        hole=0.4,
        color_discrete_sequence=px.colors.qualitative.Pastel,
        title="客户销量占比（全周期）"
    )
    fig_pie.update_traces(textposition="inside", textinfo="percent+label")
    fig_pie.update_layout(height=480, showlegend=True,
                          legend=dict(orientation="v", x=1.05))
    st.plotly_chart(fig_pie, use_container_width=True)

    # 客户排名表
    st.markdown("#### 客户销量排名")
    cust_sum["占比"] = (cust_sum["总数量"] / cust_sum["总数量"].sum() * 100).round(1).astype(str) + "%"
    st.dataframe(cust_sum, hide_index=True, use_container_width=True,
                 column_config={
                     "客户": "客户名称",
                     "总数量": st.column_config.NumberColumn("总销量（把）", format="%d"),
                     "占比": "占比"
                 })

# ─── Tab3：月度堆叠柱图 ───
with tab3:
    st.markdown("### 月度销量结构（按客户堆叠）")
    stack_data = df[df["客户"].isin(top_custs)].groupby(["月份", "客户"])["数量"].sum().reset_index()
    stack_data["月份排序"] = stack_data["月份"].map(lambda m: MONTH_ORDER.get(m, m))
    stack_data = stack_data.sort_values("月份排序")
    
    fig_stack = px.bar(
        stack_data, x="月份", y="数量", color="客户",
        barmode="stack",
        color_discrete_sequence=px.colors.qualitative.Set2,
        title=f"月度销量结构（TOP{top_n} 客户）"
    )
    fig_stack.update_layout(
        height=440, plot_bgcolor="white", paper_bgcolor="white",
        legend=dict(orientation="h", y=-0.25),
        xaxis=dict(categoryorder="array", categoryarray=months_in_data)
    )
    st.plotly_chart(fig_stack, use_container_width=True)

    # 型号维度
    st.markdown("### 型号月度分布")
    model_data = df.groupby(["月份", "型号"])["数量"].sum().reset_index()
    model_data["月份排序"] = model_data["月份"].map(lambda m: MONTH_ORDER.get(m, m))
    model_data = model_data.sort_values("月份排序")
    top_models = df.groupby("型号")["数量"].sum().nlargest(8).index.tolist()
    model_data_top = model_data[model_data["型号"].isin(top_models)]
    
    fig_model = px.bar(
        model_data_top, x="月份", y="数量", color="型号",
        barmode="group",
        color_discrete_sequence=px.colors.qualitative.Pastel1,
        title="TOP8 型号月度分布"
    )
    fig_model.update_layout(
        height=400, plot_bgcolor="white", paper_bgcolor="white",
        legend=dict(orientation="h", y=-0.25),
        xaxis=dict(categoryorder="array", categoryarray=months_in_data)
    )
    st.plotly_chart(fig_model, use_container_width=True)

# ─── Tab4：客户×月份透视表 ───
with tab4:
    st.markdown("### 客户 × 月份 销量透视表")
    pivot = df.pivot_table(
        index="客户", columns="月份", values="数量",
        aggfunc="sum", fill_value=0, margins=True, margins_name="合计"
    )
    # 按月份顺序排列列
    ordered_cols = [m for m in months_in_data if m in pivot.columns]
    if "合计" in pivot.columns:
        ordered_cols.append("合计")
    pivot = pivot[ordered_cols]
    pivot = pivot.sort_values("合计", ascending=False)

    st.dataframe(
        pivot.style.background_gradient(cmap="Blues", axis=0, subset=ordered_cols[:-1])
                   .format("{:,.0f}"),
        use_container_width=True,
        height=500
    )

    # 型号透视
    st.markdown("### 型号 × 月份 销量透视表")
    pivot_model = df.pivot_table(
        index="型号", columns="月份", values="数量",
        aggfunc="sum", fill_value=0, margins=True, margins_name="合计"
    )
    ordered_cols_m = [m for m in months_in_data if m in pivot_model.columns]
    if "合计" in pivot_model.columns:
        ordered_cols_m.append("合计")
    pivot_model = pivot_model[ordered_cols_m]
    pivot_model = pivot_model.sort_values("合计", ascending=False)
    st.dataframe(
        pivot_model.style.background_gradient(cmap="Greens", axis=0, subset=ordered_cols_m[:-1])
                         .format("{:,.0f}"),
        use_container_width=True
    )

    # 下载 Excel
    st.markdown("---")
    st.markdown("#### 下载 Excel 报表")
    excel_bytes = build_excel(
        pivot.drop(index="合计", errors="ignore"),
        df
    )
    st.download_button(
        label="⬇️ 下载 Excel 透视表",
        data=excel_bytes,
        file_name="手术刀流向分析.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ─── Tab5：明细数据 ───
with tab5:
    st.markdown("### 原始明细数据")
    col_filter1, col_filter2, col_filter3 = st.columns(3)
    with col_filter1:
        sel_months = st.multiselect("筛选月份", months_in_data, default=months_in_data)
    with col_filter2:
        all_custs = sorted(df["客户"].unique())
        sel_custs = st.multiselect("筛选客户", all_custs, default=[])
    with col_filter3:
        all_models = sorted(df["型号"].unique())
        sel_models = st.multiselect("筛选型号", all_models, default=[])
    
    df_show = df[df["月份"].isin(sel_months)]
    if sel_custs:
        df_show = df_show[df_show["客户"].isin(sel_custs)]
    if sel_models:
        df_show = df_show[df_show["型号"].isin(sel_models)]
    
    st.info(f"共 {len(df_show)} 条记录，合计 {int(df_show['数量'].sum()):,} 把")
    st.dataframe(df_show, hide_index=True, use_container_width=True,
                 column_config={
                     "数量": st.column_config.NumberColumn("数量（把）", format="%d")
                 })
    
    # CSV 下载
    csv = df_show.to_csv(index=False, encoding="utf-8-sig")
    st.download_button(
        "⬇️ 下载明细 CSV",
        data=csv.encode("utf-8-sig"),
        file_name="手术刀明细.csv",
        mime="text/csv"
    )

# ─── 底部信息 ───
st.markdown("---")
st.caption("流向数据分析工具 v1.0 | 珠海市海纳融成医疗科技有限公司 | 数据仅供内部参考")
